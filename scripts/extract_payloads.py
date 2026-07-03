#!/usr/bin/env python3
import argparse
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


def main():
    parser = argparse.ArgumentParser(description="Extract payload columns from xlsx to YAML.")
    parser.add_argument("xlsx", help="Source workbook path")
    parser.add_argument("output", help="Output YAML path")
    args = parser.parse_args()

    payloads = extract_payloads(Path(args.xlsx))
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(to_yaml(payloads, Path(args.xlsx).name), encoding="utf-8")
    total = sum(len(values) for values in payloads.values())
    print("Extracted %d categories and %d payloads to %s" % (len(payloads), total, output))


def extract_payloads(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    payloads = OrderedDict()
    seen = {}

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            continue

        headers = [cell_to_text(value) for value in header_row]
        if not any(headers):
            continue

        for header in headers:
            if header and header not in payloads:
                payloads[header] = []
                seen[header] = set()

        for row in rows:
            for index, value in enumerate(row):
                if index >= len(headers):
                    continue
                header = headers[index]
                if not header:
                    continue
                text = cell_to_text(value)
                if not text or text in seen[header]:
                    continue
                payloads[header].append(text)
                seen[header].add(text)

    return OrderedDict((key, value) for key, value in payloads.items() if value)


def cell_to_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def to_yaml(payloads, source_name):
    lines = [
        "# Generated from %s by scripts/extract_payloads.py." % source_name,
        "# Edit the workbook and rerun extraction instead of editing this file by hand.",
        "",
    ]
    for category, values in payloads.items():
        lines.append("%s:" % category)
        for value in values:
            lines.append("  - %s" % yaml_quote(value))
        lines.append("")
    return "\n".join(lines)


def yaml_quote(value):
    escaped = []
    for char in value:
        code = ord(char)
        if char == "\\":
            escaped.append("\\\\")
        elif char == "\"":
            escaped.append("\\\"")
        elif char == "\n":
            escaped.append("\\n")
        elif char == "\t":
            escaped.append("\\t")
        elif code < 0x20:
            escaped.append("\\u%04x" % code)
        else:
            escaped.append(char)
    return "\"" + "".join(escaped) + "\""


if __name__ == "__main__":
    main()

